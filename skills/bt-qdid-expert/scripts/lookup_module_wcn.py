#!/usr/bin/env python3
"""
Lookup WCN chip model from Smart Module Configuration Table XLSX.
Usage: python3 lookup_module_wcn.py <module_name>
Example: python3 lookup_module_wcn.py SC60
"""

import sys
import re
import warnings
warnings.filterwarnings("ignore")

XLSX_PATH = "/home/quectel/Documents/Smart module Configuration table_V7.1-20250603.xlsx"

def find_wcn_for_module(module_keyword: str):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    module_keyword_upper = module_keyword.strip().upper()

    results = []

    # 1. Try to find a sheet that matches the module name exactly or partially
    matching_sheets = [s for s in wb.sheetnames if module_keyword_upper in s.upper()]

    if not matching_sheets:
        # Fall back: search all sheets for the keyword in cell values
        matching_sheets = wb.sheetnames

    wcn_pattern = re.compile(r'(WCN\d+\w*)', re.IGNORECASE)
    qcs_pattern = re.compile(r'(QCS\d+\w*)', re.IGNORECASE)

    seen = set()
    found_variants = []

    for sheet_name in matching_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if cell_val is None:
                    continue
                cell_str = str(cell_val)
                # Check if this row/cell mentions our module keyword
                if module_keyword_upper not in cell_str.upper() and module_keyword_upper not in sheet_name.upper():
                    continue
                # Extract WCN / QCS chip IDs
                wcn_matches = wcn_pattern.findall(cell_str)
                qcs_matches = qcs_pattern.findall(cell_str)
                for m in wcn_matches + qcs_matches:
                    m_upper = m.upper()
                    if m_upper not in seen:
                        seen.add(m_upper)
                        found_variants.append(m_upper)

    # If we matched sheets by name, also scan entire sheet for WCN chips
    for sheet_name in matching_sheets:
        if module_keyword_upper not in sheet_name.upper():
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            chipset_col = None
            for cell_val in row:
                if cell_val is None:
                    continue
                cell_str = str(cell_val)
                wcn_matches = wcn_pattern.findall(cell_str)
                for m in wcn_matches:
                    m_upper = m.upper()
                    if m_upper not in seen:
                        seen.add(m_upper)
                        found_variants.append(m_upper)

    if found_variants:
        print(f"Module: {module_keyword}")
        print(f"Found WCN/QCS chips: {', '.join(sorted(set(found_variants)))}")
        print(f"Matching sheets: {', '.join(matching_sheets)}")
        return sorted(set(found_variants))
    else:
        print(f"No WCN/QCS chips found for module: {module_keyword}")
        print(f"Searched sheets: {', '.join(matching_sheets)}")
        return []


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 lookup_module_wcn.py <module_name>")
        print("Example: python3 lookup_module_wcn.py SC60")
        sys.exit(1)

    module = sys.argv[1]
    chips = find_wcn_for_module(module)
    if chips:
        sys.exit(0)
    else:
        sys.exit(1)
