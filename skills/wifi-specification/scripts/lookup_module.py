#!/usr/bin/env python3
"""
wifi_specification helper script.

Usage:
  # Query a module (uses local cache if available, falls back to PDF scan)
  python3 lookup_module.py <module_name> [--json]

  # Build / refresh the local knowledge cache from PDFs (run locally only)
  python3 lookup_module.py --build-cache [--qcom-dir <path>]

  # Force a live PDF scan, ignoring cache
  python3 lookup_module.py <module_name> --no-cache [--json]

PRIVACY NOTE:
  - PDF scanning is done 100% locally; raw document text never reaches stdout.
  - The cache (data/wcn_protocols.json) contains only protocol names and doc
    filenames — no confidential document text.
  - The cache file is safe to share alongside the skill; original PDFs are not.
  - Use --raw only for local debugging, never in AI-assisted workflows.
"""

import sys
import os
import re
import json
import datetime
import argparse
import warnings
warnings.filterwarnings("ignore")

XLSX_DEFAULT   = "/home/quectel/Documents/Smart module Configuration table_V7.1-20250603.xlsx"
QCOM_DIR_DEFAULT = "/home/quectel/Documents/Wifi/Qcom"
CACHE_DEFAULT  = os.path.join(os.path.dirname(__file__), "..", "data", "wcn_protocols.json")

# Chip -> subdirectory mapping (most specific match first).
# Covers both WCN and QCA naming conventions used by Qualcomm.
# Subdirectory "." means the root QCOM_DIR (used when no dedicated subdir exists).
WCN_DIR_MAP = [
    (re.compile(r'WCN785[0-9]', re.IGNORECASE), ["78xx"]),
    (re.compile(r'WCN78',       re.IGNORECASE), ["78xx"]),
    (re.compile(r'WCN6856',     re.IGNORECASE), ["WCN6856"]),
    (re.compile(r'WCN685',      re.IGNORECASE), ["WCN6856"]),
    (re.compile(r'WCN675',      re.IGNORECASE), ["WCN6856"]),
    (re.compile(r'WCN3988',     re.IGNORECASE), ["3988", "39xx"]),
    (re.compile(r'WCN3980',     re.IGNORECASE), ["39xx"]),
    (re.compile(r'WCN399',      re.IGNORECASE), ["39xx"]),
    (re.compile(r'WCN398',      re.IGNORECASE), ["39xx"]),
    (re.compile(r'WCN3950',     re.IGNORECASE), ["3950", "39xx"]),
    (re.compile(r'WCN39',       re.IGNORECASE), ["39xx"]),
    (re.compile(r'WCN368',      re.IGNORECASE), ["36xx"]),
    (re.compile(r'WCN366',      re.IGNORECASE), ["36xx"]),
    (re.compile(r'WCN36',       re.IGNORECASE), ["36xx"]),
    # QCA series — docs live in the root Qcom dir (no dedicated subdir)
    (re.compile(r'QCA639',      re.IGNORECASE), ["."]),   # QCA6391, QCA6390 …
    (re.compile(r'QCA102',      re.IGNORECASE), ["."]),   # QCA1023 …
    (re.compile(r'QCA106',      re.IGNORECASE), ["."]),   # QCA1064 …
    (re.compile(r'QCA',         re.IGNORECASE), ["."]),   # fallback for other QCA chips
]

# One representative chip name per known family, used when building the cache
WCN_FAMILIES = [
    "WCN3660",
    "WCN3950",
    "WCN3988",
    "WCN6856",
    "WCN7850",
    # QCA series (WiFi-only chips, docs in root Qcom dir)
    "QCA6391",
    "QCA1023",
]

SECURITY_RE = re.compile(
    r'(WPA3|WPA2|WPA[\s\-/]?[12]?[\s\-]?(Personal|Enterprise|Mixed)?'
    r'|WEP|TKIP|CCMP|GCMP|AES[\s\-]?CCMP|SAE|OWE|Enhanced Open'
    r'|Suite[\s\-]?B|192[\s\-]?bit|802\.11[rwi]|WAPI|EAP[\s\-][A-Z]+'
    r'|PMF|PMKSA|WPS[\s\d.]*|Passpoint|DPP|FILS)',
    re.IGNORECASE
)

# Protocol detection rules: (label, [regex patterns])
PROTOCOL_RULES = [
    ("WPA3-Personal (SAE)",
     [r'WPA3[\s\-]Personal', r'\bSAE\b', r'isWpa3SaeSupported']),
    ("WPA3-Enterprise Suite-B-192",
     [r'WPA3[\s\-]Enterprise', r'Suite[\s\-]?B[\s\-]?192', r'SUITE_B', r'isWpa3SuiteBSupported',
      r'CONFIG_SUITEB192', r'SUITEB192', r'SUITEB', r'suite_b_192',
      r'AKM.*13\b', r'AKM.*12\b',
      r'wpa3.*enterprise', r'enterprise.*wpa3',
      r'EAP-TLS.*256', r'GCMP[\s\-]256.*enterprise', r'AES[\s\-]256[\s\-]CMAC']),
    ("WPA3 Enhanced Open (OWE)",
     [r'\bOWE\b', r'Enhanced Open', r'Opportunistic Wireless']),
    ("WPA2-Personal (CCMP/AES)",
     [r'\bWPA2\b', r'WPA[-/ ]?2\b']),
    ("WPA2-Enterprise (EAP/802.1X)",
     [r'WPA2.*Enterprise', r'WPA.*Enterprise', r'EAP[\s\-](TLS|TTLS|PEAP|AKA|SIM)',
      r'802\.1X', r'RADIUS']),
    ("WPA-Personal (TKIP)",
     [r'\bWPA\b(?![\s\-]?[23])', r'WPA[-/ ]?Personal', r'\bTKIP\b']),
    ("WEP",
     [r'\bWEP\b']),
    ("WAPI",
     [r'\bWAPI\b']),
    ("PMF / 802.11w (Protected Management Frames)",
     [r'\bPMF\b', r'802\.11w', r'GMAC']),
    ("WPS 2.0",
     [r'\bWPS[\s\d.]*']),
    ("Passpoint (Hotspot 2.0)",
     [r'\bPasspoint\b', r'Hotspot 2\.0', r'HS2']),
    ("DPP (Device Provisioning Protocol)",
     [r'\bDPP\b', r'Device Provisioning Protocol']),
    ("FILS (Fast Initial Link Setup)",
     [r'\bFILS\b', r'Fast Initial Link']),
    ("GCMP-256",
     [r'\bGCMP\b', r'GCMP-256']),
]

# Regex to select feature-relevant lines from PDFs
FEATURE_RE = re.compile(
    r'(MU.?MIMO|SU.?MIMO|OFDMA|TWT|Target Wake Time|[Bb]eamform'
    r'|802\.11[acgnrkvwi]+'
    r'|2\.4\s*GHz|5\s*GHz|6\s*GHz'
    r'|HT20\b|HT40|VHT20\b|VHT40\b|VHT80|40\s*MHz|80\s*MHz|160\s*MHz|320\s*MHz|80\s*\+\s*80'
    r'|DBS|SBS|DBDC|Dual.Band.Simult'
    r'|[Rr]oam|LFR\b|LFR3|11[rkv]\b|Fast BSS|FT.SAE|FT.PSK|MBO\b|OCE\b|PER.based'
    r'|[Rr]oam.*[Oo]ffload|[Oo]ffload.*[Rr]oam'
    r'|P2P|Wi.?Fi Direct|SoftAP|Soft AP|STA.SAP|Wi.Fi [Ss]har'
    r'|\bNAN\b|\bRTT\b|\bFTM\b|802\.11mc'
    r'|\bTDLS\b|Tunneled Direct'
    r'|WMM|APSD|U.APSD|UAPSD|QoS|WMM.PS|WMM.AC|[Vv]oice [Ee]nterprise|CCXv4'
    r'|\bIMPS\b|\bBMPS\b|WoW\b|Wake.on.Wireless|QPower|\bPNO\b|ARP.*[Oo]ffload'
    r'|\bGSCAN\b|[Gg]eofence|[Ll]ow.power.*scan|[Ss]can.*[Oo]ffload|[Hh]otspot.*[Oo]ffload'
    r'|[Cc]ontext [Hh]ub'
    r'|GreenTx|Green Tx|[Ee]nhanced [Gg]reen|Green.*AP'
    r'|\bSAR\b|SAR.*backoff|[Tt]x.*power.*limit|[Pp]ower.*backoff'
    r'|\bDFS\b|Dynamic Freq|\bACS\b|Auto.?[Cc]hannel'
    r'|\bMCC\b|\bSCC\b|Multi.?[Cc]hannel [Cc]oncurr'
    r'|Bluetooth.*[Cc]oex|BT.*[Cc]oex|\bBTC\b|LTE.*[Cc]oex|LTE.WLAN|TWS.*[Cc]oex'
    r'|LTE\s*\+\s*LTE|LTE.*LTE'
    r'|[Aa]ntenna.shar|WAN ASM'
    r'|[Hh]idden.SSID|[Bb]lacklist|[Ww]hitelist|[Kk]ick.out|[Cc]lient.*limit|max.*client'
    r'|[Mm]ac.*[Rr]andom|[Rr]andom.*MAC|[Mm]ac.*[Ss]poof|[Pp]rivacy.*[Ss]ecurity)',
    re.IGNORECASE
)

# Feature detection rules: category -> [(label, [patterns])]
FEATURE_RULES = {
    "rf_features": [
        ("OFDMA",                        [r'\bOFDMA\b']),
        ("MU-MIMO",                      [r'\bMU.?MIMO\b']),
        ("SU-MIMO / 2×2 MIMO",           [r'\bSU.?MIMO\b', r'2\s*[×xX]\s*2\s*MIMO', r'\b2x2\b']),
        ("4×4 MIMO / 8×8 MIMO",          [r'4\s*[×xX]\s*4\s*MIMO', r'8\s*[×xX]\s*8', r'\b4x4\b']),
        ("Beamforming (Tx/Rx)",          [r'[Bb]eamform', r'[Tt]ransmit [Bb]eamform']),
        ("TWT (Target Wake Time)",        [r'\bTWT\b', r'Target Wake Time']),
        ("GreenTx / Enhanced Green AP",  [r'GreenTx', r'Green Tx', r'[Ee]nhanced [Gg]reen', r'Green.*AP']),
        ("SAR / Tx Power Backoff",       [r'\bSAR\b', r'SAR.*backoff', r'[Tt]x.*power.*limit',
                                          r'[Pp]ower.*backoff', r'[Pp]ower.*SAR']),
    ],
    "channel_and_phy": [
        ("802.11a/b/g (legacy modes)",   [r'802\.11[abg]\b', r'802\.11a/b/g']),
        ("802.11n (HT20/HT40)",          [r'802\.11n', r'\bHT20\b', r'\bHT40\b']),
        ("802.11ac (VHT20/40/80)",       [r'802\.11ac', r'\bVHT20\b', r'\bVHT40\b', r'\bVHT80\b']),
        ("802.11ax (HE / Wi-Fi 6/6E)",   [r'802\.11ax', r'\bHE\b', r'Wi.?Fi 6']),
        ("320 MHz (EHT)",                [r'320\s*MHz']),
        ("160 MHz / 80+80 MHz",          [r'160\s*MHz', r'80\s*\+\s*80']),
        ("80 MHz (VHT80)",               [r'VHT80', r'(?<!\+\s)80\s*MHz']),
        ("40 MHz (HT40)",                [r'HT40', r'40\s*MHz']),
        ("DFS (Dynamic Frequency Selection)", [r'\bDFS\b', r'Dynamic Freq']),
        ("ACS (Auto Channel Selection)", [r'\bACS\b', r'Auto.?[Cc]hannel [Ss]elect',
                                          r'Automatic.*[Cc]hannel.*[Ss]elect']),
    ],
    "qos": [
        ("WMM (Wi-Fi Multimedia)",       [r'\bWMM\b']),
        ("WMM-PS / U-APSD",              [r'WMM.PS', r'\bU.APSD\b', r'\bUAPSD\b', r'\bAPSD\b']),
        ("WMM-AC (Admission Control)",   [r'WMM.AC\b']),
        ("Voice Enterprise / CCXv4",     [r'[Vv]oice [Ee]nterprise', r'\bCCX[Vv]?4\b']),
        ("QoS (General)",                [r'\bQoS\b']),
    ],
    "roaming": [
        ("802.11r (Fast BSS Transition)", [r'802\.11r', r'Fast BSS', r'FT.SAE', r'FT.PSK']),
        ("802.11k (Neighbor Report)",    [r'802\.11k', r'\b11k\b', r'Neighbor Report']),
        ("802.11v (BSS Transition Mgmt)",[r'802\.11v', r'\b11v\b', r'BSS Transition']),
        ("LFR / LFR3.0 (Fast Roaming)", [r'\bLFR\b', r'LFR3', r'Legacy Fast Roam',
                                          r'[Rr]oam.*[Oo]ffload', r'[Oo]ffload.*[Rr]oam']),
        ("MBO (Multiband Operations)",   [r'\bMBO\b', r'[Mm]ultiband [Oo]peration']),
        ("OCE (Optimized Connectivity)", [r'\bOCE\b', r'Optimized Connectivity']),
        ("PER-based Roaming",            [r'PER.based [Rr]oam']),
    ],
    "location": [
        ("RTT / FTM (802.11mc ranging)", [r'\bRTT\b', r'\bFTM\b', r'802\.11mc',
                                          r'[Rr]anging', r'Fine Timing']),
        ("GSCAN / Geofencing",           [r'\bGSCAN\b', r'[Gg]eofence', r'[Gg]eofencing']),
        ("Low-power Wi-Fi location scan",[r'[Ll]ow.power.*scan', r'[Ss]can.*[Oo]ffload',
                                          r'location.*scan', r'[Bb]ackground.*scan']),
        ("Context Hub support",          [r'[Cc]ontext [Hh]ub']),
    ],
    "power_save": [
        ("IMPS / BMPS",                  [r'\bIMPS\b', r'\bBMPS\b']),
        ("WoW (Wake on Wireless)",       [r'\bWoW\b', r'Wake.on.Wireless', r'Wake.on.WLAN']),
        ("D-APSD / QPower",              [r'\bQPower\b', r'D.APSD']),
        ("PNO (Preferred Network Offload)", [r'\bPNO\b', r'Preferred Network Offload']),
        ("Hotspot Offload",              [r'[Hh]otspot.*[Oo]ffload', r'[Oo]ffload.*[Hh]otspot']),
        ("ARP / NS Offload",             [r'ARP.*[Oo]ffload', r'NS.*[Oo]ffload',
                                          r'[Nn]eighbor [Dd]iscovery [Oo]ffload']),
    ],
    "concurrent_modes": [
        ("DBS (Dual Band Simultaneous)", [r'\bDBS\b', r'Dual.Band.Simult', r'\bDBDC\b']),
        ("SBS (Single Band Simultaneous)",[r'\bSBS\b']),
        ("STA+SAP / Wi-Fi Sharing",      [r'STA.SAP', r'Wi.Fi [Ss]haring', r'Wi.Fi [Ss]hare']),
        ("SoftAP",                       [r'\bSoftAP\b', r'Soft AP']),
        ("P2P / Wi-Fi Direct",           [r'\bP2P\b', r'Wi.?Fi Direct']),
        ("TDLS (Tunneled Direct Link Setup)", [r'\bTDLS\b', r'Tunneled Direct Link']),
        ("NAN (Neighbor Awareness Networking)", [r'\bNAN\b', r'Neighbor Aware']),
    ],
    "ap_mode": [
        ("Hidden SSID",                  [r'[Hh]idden.SSID', r'[Hh]idden.SSID']),
        ("STA Blacklist / Whitelist",    [r'[Bb]lacklist', r'[Ww]hitelist', r'[Kk]ick.out',
                                          r'[Cc]lient.*[Kk]ick']),
        ("Max 32 STA clients",           [r'32.*[Cc]lient', r'[Cc]lient.*limit.*32',
                                          r'[Nn]umber.*[Cc]lient.*32']),
        ("SAP ACS",                      [r'SAP.*ACS', r'[Ss]oft.?AP.*[Cc]hannel.*[Ss]elect']),
    ],
    "privacy": [
        ("MAC Randomization",            [r'[Mm]ac.*[Rr]andom', r'[Rr]andom.*MAC',
                                          r'[Rr]andom.*[Mm]ac [Aa]ddress']),
        ("MAC Address Spoofing",         [r'[Mm]ac.*[Ss]poof', r'MAC.*[Ss]poofing',
                                          r'WLAN MAC.*[Ss]poof']),
        ("Privacy Security Features",    [r'[Pp]rivacy.*[Ss]ecurity', r'[Pp]rivacy.*[Ff]eature']),
    ],
    "coexistence": [
        ("Bluetooth Coexistence",        [r'Bluetooth.*[Cc]oex', r'BT.*[Cc]oex', r'BTC\b']),
        ("LTE / Cellular Coexistence",   [r'LTE.*[Cc]oex', r'LTE.WLAN', r'LTE.LAA', r'LTE.U\b']),
        ("LTE + LTE Coexistence",        [r'LTE\s*\+\s*LTE', r'LTE.*LTE.*[Cc]oex']),
        ("TWS Coexistence",              [r'TWS.*[Cc]oex', r'[Cc]oex.*TWS']),
        ("Antenna Sharing (LTE/WAN)",    [r'[Aa]ntenna.shar', r'WAN ASM', r'LTE.*antenna']),
        ("MCC (Multi-Channel Concurrency)", [r'\bMCC\b', r'Multi.?[Cc]hannel [Cc]oncurr']),
        ("SCC (Single-Channel Concurrency)", [r'\bSCC\b', r'Single.?[Cc]hannel [Cc]oncurr']),
    ],
}


def detect_features(snippets):
    """Detect WiFi features from PDF snippets; return dict of category -> [labels]."""
    joined = " ".join(snippets)
    result = {}
    for category, rules in FEATURE_RULES.items():
        found = [label for label, patterns in rules
                 if any(re.search(p, joined, re.IGNORECASE) for p in patterns)]
        if found:
            result[category] = found
    return result


# ── xlsx WiFi band extraction ─────────────────────────────────────────────────

def _wifi_standard_label(raw):
    """Derive a human-readable WiFi generation label from a raw band string."""
    raw_up = raw.upper()
    if re.search(r'802\.11AX|AX\b|WI-FI\s*6', raw_up):
        if re.search(r'6\s*GHZ', raw_up):
            return "WiFi 6E (802.11ax, 2.4/5/6 GHz)"
        return "WiFi 6 (802.11ax, 2.4/5 GHz)"
    if re.search(r'802\.11AC|AC\b', raw_up):
        return "WiFi 5 (802.11ac, 2.4/5 GHz)"
    if re.search(r'802\.11N', raw_up):
        return "WiFi 4 (802.11n)"
    return raw.strip()


def extract_doc_version(pdf_path):
    """Extract document number, revision and date from a PDF cover/revision page.

    Returns a dict with keys: doc_number, revision, date, title.
    All values are strings or None. Never returns raw document text.
    """
    try:
        import pdfplumber
    except ImportError:
        return {}

    DOC_NUM_RE  = re.compile(r'(80-[A-Z0-9]+-\d+)\s+Rev\.?\s+([A-Z]+)', re.IGNORECASE)
    DATE_RE     = re.compile(r'(January|February|March|April|May|June|July|August|September'
                              r'|October|November|December)\s+\d+,?\s+20\d\d'
                              r'|\b20\d\d-\d\d-\d\d\b', re.IGNORECASE)
    TITLE_RE    = re.compile(r'(WLAN|WiFi|Wi-Fi|WCN\d+|QCA\d+|Wireless|Connectivity|Software)'
                              r'.{5,80}(Guide|Overview|Manual|Report|Document)', re.IGNORECASE)

    result = {"doc_number": None, "revision": None, "date": None, "title": None}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for pg_idx in range(min(3, len(pdf.pages))):
                text = pdf.pages[pg_idx].extract_text() or ""
                for line in text.split("\n"):
                    line = line.strip()
                    if result["doc_number"] is None:
                        m = DOC_NUM_RE.search(line)
                        if m:
                            result["doc_number"] = m.group(1)
                            result["revision"]   = m.group(2).upper()
                    if result["date"] is None:
                        m = DATE_RE.search(line)
                        if m:
                            result["date"] = m.group(0).strip()
                    if result["title"] is None and len(line) > 15:
                        m = TITLE_RE.search(line)
                        if m:
                            result["title"] = re.sub(r'\s+', ' ', line).strip()[:120]
                if result["doc_number"] and result["date"]:
                    break
    except Exception:
        pass
    return {k: v for k, v in result.items() if v is not None}


def extract_all_module_platforms(xlsx_path):
    """Scan every sheet and return comprehensive per-variant platform data.

    Returns:
        {
          "SC200U": {
            "wifi_standard": "WiFi 5 (802.11ac, 2.4/5 GHz)",
            "frequency_bands": ["2.4 GHz", "5 GHz"],
            "variants": [
              {
                "name": "SC200U-EMNA",
                "chipset": "SM4250+PM4250+WTR-2965+...",
                "wifi_chips": ["WCN3950"],
                "pcb": "V1.1",
                "prx_band": "FDD-LTE:B1/2/3...",
                "drx_band": "...",
                "mimo": null,
                "wifi_band": "802.11 a/b/g/n/ac",
                "description": "DDR+EMMC (4GB+64GB)",
                "notes": "..."
              }, ...
            ]
          }, ...
        }
    """
    import openpyxl
    wcn_re = re.compile(r'(?:WCN|QCA)[-]?\d+[A-Z\d]*', re.IGNORECASE)

    # Column header keywords (case-insensitive substring match)
    COL_HINTS = {
        "variant":     [r'版本',    r'[Vv]ariant', r'^型号$'],
        "chipset":     [r'[Cc]hipset', r'[Cc]hip\s*\+'],
        "pcb":         [r'\bPCB\b'],
        "prx_band":    [r'PRX',    r'[Pp]rimary.*[Bb]and'],
        "drx_band":    [r'DRX',    r'[Dd]iversity.*[Bb]and'],
        "mimo":        [r'\bMimo\b', r'\bMIMO\b'],
        "wifi_band":   [r'Wi.?Fi\s*band', r'WLAN\s*band'],
        "description": [r'说明',   r'[Dd]escription', r'[Mm]emory'],
        "notes":       [r'备注',   r'[Nn]otes?\b',    r'[Rr]emark'],
    }

    def detect_cols(ws):
        """Return {field: col_index} by scanning header rows."""
        cols = {}
        for hrow in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            if not hrow or not any(hrow):
                continue
            for idx, cell in enumerate(hrow):
                if cell is None:
                    continue
                cell_str = str(cell)
                for field, patterns in COL_HINTS.items():
                    if field not in cols:
                        if any(re.search(p, cell_str) for p in patterns):
                            cols[field] = idx
            if "chipset" in cols and "wifi_band" in cols:
                break
        return cols

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols, header_row = {}, 1  # header_row tracks which row the header was found

        for h_idx, hrow in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), start=1):
            if not hrow or not any(hrow):
                continue
            for idx, cell in enumerate(hrow):
                if cell is None:
                    continue
                cell_str = str(cell)
                for field, patterns in COL_HINTS.items():
                    if field not in cols:
                        if any(re.search(p, cell_str) for p in patterns):
                            cols[field] = idx
            if "chipset" in cols and "wifi_band" in cols:
                header_row = h_idx
                break

        if "chipset" not in cols:
            continue  # sheet has no chipset column — skip

        family = re.split(r'[&（(]', sheet_name)[0].strip()
        variants = []
        best_wifi = ""

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row:
                continue

            def cell(field):
                idx = cols.get(field)
                if idx is None or len(row) <= idx:
                    return None
                v = row[idx]
                if v is None:
                    return None
                s = str(v).replace('\r\n', ' ').replace('\n', ' ').strip()
                return s if s not in ('/', 'None', '') else None

            chipset = cell("chipset")
            if not chipset:
                continue  # skip rows without chipset info

            variant_name = cell("variant") or family
            wifi_raw     = cell("wifi_band") or ""
            pcb          = cell("pcb")
            prx          = cell("prx_band")
            drx          = cell("drx_band")
            mimo_raw     = cell("mimo")
            desc         = cell("description")
            notes        = cell("notes")

            # Extract WiFi/WCN chips from chipset string
            wifi_chips = list(dict.fromkeys(
                re.sub(r'(WCN|QCA)(\d)', r'\1\2', m.upper().replace('-', ''))
                for m in wcn_re.findall(chipset)
            ))

            # Track best (longest) wifi_band string for wifi_standard label
            if len(wifi_raw) > len(best_wifi):
                best_wifi = wifi_raw

            entry = {
                "name":        variant_name,
                "chipset":     chipset[:120],
                "wifi_chips":  wifi_chips,
                "pcb":         pcb,
                "prx_band":    prx[:200] if prx else None,
                "drx_band":    drx[:200] if drx else None,
                "mimo":        mimo_raw[:80] if mimo_raw else None,
                "wifi_band":   wifi_raw[:80] if wifi_raw else None,
                "description": desc[:100] if desc else None,
                "notes":       notes[:100] if notes else None,
            }
            # Remove None values to keep JSON compact
            entry = {k: v for k, v in entry.items() if v is not None}
            variants.append(entry)

        if not variants:
            continue

        # Deduplicate variants (same name+chipset+pcb = same variant)
        seen_keys, deduped = set(), []
        for v in variants:
            key = (v.get("name", ""), v.get("chipset", ""), v.get("pcb", ""))
            if key not in seen_keys:
                seen_keys.add(key)
                deduped.append(v)

        # Derive frequency bands from best wifi string
        bands = []
        for ghz in ["6 GHz", "5 GHz", "2.4 GHz"]:
            if re.search(ghz.replace(" ", r"\s*").replace(".", r"\."), best_wifi, re.IGNORECASE):
                bands.append(ghz)

        results[family] = {
            "wifi_standard":    _wifi_standard_label(best_wifi) if best_wifi else None,
            "frequency_bands":  bands,
            "variant_count":    len(deduped),
            "variants":         deduped,
        }

    return results


# Keep backward-compat alias (used by older cache query paths)
def extract_all_module_wifi(xlsx_path):
    platforms = extract_all_module_platforms(xlsx_path)
    # Return simplified wifi-only dict for any old callers
    return {
        family: {
            "wifi_standard":   data.get("wifi_standard"),
            "frequency_bands": data.get("frequency_bands", []),
            "variant_count":   data.get("variant_count", 0),
        }
        for family, data in platforms.items()
    }


# ── Cache helpers ─────────────────────────────────────────────────────────────

def load_cache(cache_path):
    """Load the local knowledge cache. Returns {} if not found."""
    path = os.path.abspath(cache_path)
    if os.path.isfile(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_cache(data, cache_path):
    """Save the knowledge cache to disk."""
    path = os.path.abspath(cache_path)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def canonical_chip_key(wcn_chip):
    """Map any WCN chip string to a canonical family key for cache lookup."""
    for pattern, _ in WCN_DIR_MAP:
        if pattern.search(wcn_chip):
            # Derive key from the first matching family representative
            for fam in WCN_FAMILIES:
                if pattern.search(fam):
                    return fam.upper()
            # Fallback: use the chip itself normalised
            return re.sub(r'[-\s]', '', wcn_chip).upper()
    return wcn_chip.upper()


def protocols_from_cache(wcn_chips, cache):
    """Return (protocols, features, source_docs, cache_hit) from the local cache."""
    chips_data = cache.get("chips", {})
    all_protocols, all_docs = set(), []
    all_features: dict = {}
    any_hit = False
    for chip in wcn_chips:
        key = canonical_chip_key(chip)
        if key in chips_data:
            any_hit = True
            all_protocols.update(chips_data[key].get("protocols", []))
            all_docs.extend(chips_data[key].get("source_docs", []))
            for cat, items in chips_data[key].get("features", {}).items():
                all_features.setdefault(cat, [])
                for item in items:
                    if item not in all_features[cat]:
                        all_features[cat].append(item)
    return sorted(all_protocols), all_features, list(dict.fromkeys(all_docs)), any_hit


# ── Core extraction ───────────────────────────────────────────────────────────

def find_module_wcn(module_query, xlsx_path):
    """Return list of (sheet_name, module_variant, [wcn_chips]) matches."""
    import openpyxl
    # Match both WCN and QCA chip naming conventions
    wcn_pattern = re.compile(r'(?:WCN|QCA)[-]?\d+[A-Z\d]*', re.IGNORECASE)
    query = module_query.strip().upper()
    module_family = re.split(r'[-_]', query)[0]
    results = []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet_upper = sheet_name.upper()
        if module_family not in sheet_upper and query not in sheet_upper:
            if not any(part in sheet_upper for part in query.split('-') if len(part) > 2):
                continue

        ws = wb[sheet_name]

        # Auto-detect column layout from header row
        variant_col, chipset_col = None, None
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if not row:
                continue
            for idx, cell in enumerate(row):
                if cell and 'Chipset' in str(cell):
                    chipset_col = idx
                    variant_col = max(0, idx - 1)
                    break
            if chipset_col is not None:
                break

        candidate_pairs = [(variant_col, chipset_col)] if chipset_col is not None \
                          else [(0, 1), (1, 2)]

        for v_col, c_col in candidate_pairs:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= c_col:
                    continue
                variant_cell = row[v_col]
                chipset_cell = row[c_col]
                if not variant_cell or not chipset_cell:
                    continue
                variant_str = str(variant_cell).strip()
                variant_upper = variant_str.upper()
                chipset_str = str(chipset_cell)
                if query in variant_upper or variant_upper.startswith(query) or variant_upper == query:
                    wcns = list(set(
                        re.sub(r'WCN(\d)', r'WCN\1', m.upper().replace('-', ''))
                        for m in wcn_pattern.findall(chipset_str)
                    ))
                    if wcns:
                        results.append((sheet_name, variant_str, wcns))
    return results


def find_qcom_docs(wcn_chips, qcom_dir):
    """Return list of PDF paths relevant to the given chips (WCN or QCA)."""
    found_dirs = set()
    # Collect chips that map to root dir "." — we'll filter those by filename
    root_chips = []
    for chip in wcn_chips:
        for pattern, subdirs in WCN_DIR_MAP:
            if pattern.search(chip):
                for sd in subdirs:
                    found_dirs.add(sd)
                if "." in subdirs:
                    root_chips.append(chip)
                break
    if not found_dirs:
        found_dirs = {"."}

    # Build a filename keyword filter for root-dir QCA chips so we don't scan
    # every unrelated PDF in the root Qcom directory.
    # Use 5-char prefix (e.g. "qca63" for QCA6391, "qca10" for QCA1023/1064)
    # so sibling chip docs in the same family are also matched.
    root_keywords = []
    for chip in root_chips:
        root_keywords.append(chip[:5].lower())

    pdf_paths = []
    for d in sorted(found_dirs):
        dirpath = os.path.join(qcom_dir, d)
        if not os.path.isdir(dirpath):
            continue
        for f in sorted(os.listdir(dirpath)):
            if not f.lower().endswith(".pdf"):
                continue
            # For root dir: only include PDFs whose filename hints at the chip family
            if d == "." and root_keywords:
                flower = f.lower()
                if not any(kw in flower for kw in root_keywords):
                    continue
            pdf_paths.append(os.path.join(dirpath, f))
    return pdf_paths


def extract_from_pdf(pdf_path):
    """Extract security + feature relevant lines from a PDF (local only, never to stdout)."""
    try:
        import pdfplumber
    except ImportError:
        return [], []
    sec_snippets, feat_snippets = [], []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    line = line.strip()
                    if len(line) > 8:
                        if SECURITY_RE.search(line):
                            sec_snippets.append(line)
                        if FEATURE_RE.search(line):
                            feat_snippets.append(line)
    except Exception as e:
        sec_snippets.append(f"[Error reading PDF: {e}]")

    def dedup(lst):
        seen, out = set(), []
        for s in lst:
            k = re.sub(r'\s+', ' ', s).lower()
            if k not in seen:
                seen.add(k)
                out.append(s)
        return out

    return dedup(sec_snippets), dedup(feat_snippets)


# Keep backward-compat alias used by --no-cache path
def extract_security_from_pdf(pdf_path):
    sec, _ = extract_from_pdf(pdf_path)
    return sec


def detect_protocols(snippets):
    """Match snippets against protocol rules; return sorted list of protocol labels."""
    joined = " ".join(snippets)
    return [label for label, patterns in PROTOCOL_RULES
            if any(re.search(p, joined, re.IGNORECASE) for p in patterns)]


def scan_chip_data(chip_name, qcom_dir):
    """Scan PDFs for a chip; return (protocols, features, doc_names, hit_counts, doc_versions)."""
    doc_paths = find_qcom_docs([chip_name], qcom_dir)
    all_sec, all_feat, hit_counts, doc_versions = [], [], {}, {}
    for pdf_path in doc_paths:
        sec, feat = extract_from_pdf(pdf_path)
        name = os.path.basename(pdf_path)
        hit_counts[name] = len(sec)
        all_sec.extend(sec)
        all_feat.extend(feat)
        ver = extract_doc_version(pdf_path)
        if ver:
            doc_versions[name] = ver
    protocols = detect_protocols(all_sec)
    features  = detect_features(all_feat)
    doc_names = [os.path.basename(p) for p in doc_paths]
    return protocols, features, doc_names, hit_counts, doc_versions


def scan_chip_protocols(chip_name, qcom_dir):
    """Backward-compat wrapper."""
    protocols, _, doc_names, hit_counts, _ = scan_chip_data(chip_name, qcom_dir)
    return protocols, doc_names, hit_counts


# ── Cache build ───────────────────────────────────────────────────────────────

def cmd_build_cache(qcom_dir, xlsx_path, cache_path):
    """Scan all chip PDFs + full xlsx platform data and write the local knowledge cache (v3.0)."""
    print(f"\nBuilding local knowledge cache  [v3.0]")
    print(f"  PDF source : {qcom_dir}")
    print(f"  xlsx source: {xlsx_path}")
    print(f"  Output     : {os.path.abspath(cache_path)}\n")

    # 1. Chip-level data from PDFs (protocols + features + doc versions)
    chips_data = {}
    for chip in WCN_FAMILIES:
        key = chip.upper()
        print(f"  [chip] Scanning {key} ...", end=" ", flush=True)
        protocols, features, doc_names, hit_counts, doc_versions = scan_chip_data(chip, qcom_dir)
        chips_data[key] = {
            "protocols":     protocols,
            "features":      features,
            "source_docs":   doc_names,
            "doc_hit_counts": hit_counts,
            "doc_versions":  doc_versions,
        }
        feat_count = sum(len(v) for v in features.values())
        ver_count  = len(doc_versions)
        print(f"{len(protocols)} protocols, {feat_count} features, "
              f"{len(doc_names)} docs ({ver_count} with version info)")

    # 2. Module-level platform info from xlsx (full variant data)
    print(f"\n  [xlsx] Extracting full module platform info ...", end=" ", flush=True)
    modules_data = extract_all_module_platforms(xlsx_path)
    total_variants = sum(d.get("variant_count", 0) for d in modules_data.values())
    print(f"{len(modules_data)} module families, {total_variants} total variants")

    cache = {
        "_meta": {
            "built_at":      datetime.datetime.now().isoformat(timespec="seconds"),
            "qcom_dir":      qcom_dir,
            "xlsx_path":     xlsx_path,
            "chip_families": WCN_FAMILIES,
            "version":       "3.0",
            "note": ("Distilled from Qualcomm confidential documents and module configuration xlsx. "
                     "Contains only protocol/feature names, WiFi specs and variant platform info — "
                     "no document text. Safe to distribute alongside the skill."),
        },
        "chips":   chips_data,
        "modules": modules_data,
    }
    save_cache(cache, cache_path)

    total_protocols = sum(len(v['protocols']) for v in chips_data.values())
    total_features  = sum(sum(len(f) for f in v['features'].values()) for v in chips_data.values())
    total_doc_vers  = sum(len(v.get('doc_versions', {})) for v in chips_data.values())
    print(f"\n✔ Cache written: {os.path.abspath(cache_path)}")
    print(f"  Chips         : {', '.join(WCN_FAMILIES)}")
    print(f"  Protocols     : {total_protocols} detections across all chip families")
    print(f"  Features      : {total_features} detections across all chip families")
    print(f"  Doc versions  : {total_doc_vers} documents with version metadata")
    print(f"  Module families: {len(modules_data)} ({total_variants} variant entries)\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="WiFi security protocol lookup for Quectel modules")
    parser.add_argument("module", nargs="?",
                        help="Module name (e.g. SC665S, SC200U, SG560D). Omit when using --build-cache.")
    parser.add_argument("--xlsx",      default=XLSX_DEFAULT)
    parser.add_argument("--qcom-dir",  default=QCOM_DIR_DEFAULT)
    parser.add_argument("--cache",     default=CACHE_DEFAULT, help="Path to local knowledge cache JSON")
    parser.add_argument("--build-cache", action="store_true",
                        help="Scan all WCN chip PDFs and build/refresh the local knowledge cache")
    parser.add_argument("--no-cache",  action="store_true",
                        help="Ignore cache and always scan PDFs directly")
    parser.add_argument("--json",      action="store_true",
                        help="Output machine-readable JSON (no raw document text)")
    parser.add_argument("--raw",       action="store_true",
                        help="[DEBUG ONLY] Print raw PDF snippets — do NOT use in AI workflows")
    args = parser.parse_args()

    # ── Build-cache mode ─────────────────────────────────────────────────────
    if args.build_cache:
        cmd_build_cache(args.qcom_dir, args.xlsx, args.cache)
        sys.exit(0)

    if not args.module:
        parser.error("module name is required (or use --build-cache)")

    # ── Step 1: xlsx lookup ──────────────────────────────────────────────────
    matches = find_module_wcn(args.module, args.xlsx)
    if not matches:
        err = {"error": f"No entries found for '{args.module}'", "module": args.module}
        print(json.dumps(err) if args.json else
              f"[ERROR] No entries found for '{args.module}'.\n  Tip: Try a shorter prefix.")
        sys.exit(1)

    all_wcns = set()
    variants = []
    for sheet, variant, wcns in matches:
        all_wcns.update(wcns)
        variants.append(variant)

    # ── Step 2: resolve protocols + features (cache-first) ───────────────────
    cache_used = False
    features, module_wifi = {}, {}
    if not args.no_cache:
        cache = load_cache(args.cache)
        if cache:
            protocols, features, doc_names, cache_used = protocols_from_cache(list(all_wcns), cache)
            doc_hit_counts = {}
            # Module-level info: look up the query prefix in the modules section
            query_family = re.split(r'[-_]', args.module.upper())[0]
            for key, info in cache.get("modules", {}).items():
                if key.upper() == query_family or key.upper().startswith(query_family):
                    module_wifi = info
                    break

    if not cache_used:
        # Fall back to live PDF scan
        doc_paths    = find_qcom_docs(list(all_wcns), args.qcom_dir)
        doc_names    = [os.path.basename(p) for p in doc_paths]
        all_sec, all_feat, doc_hit_counts = [], [], {}
        for pdf_path in doc_paths:
            sec, feat = extract_from_pdf(pdf_path)
            doc_hit_counts[os.path.basename(pdf_path)] = len(sec)
            all_sec.extend(sec)
            all_feat.extend(feat)
            if args.raw:
                print(f"\n--- [RAW] {os.path.basename(pdf_path)} ({len(sec)} lines) ---")
                for s in sec[:20]:
                    print(f"  {s[:120]}")
        protocols = detect_protocols(all_sec)
        features  = detect_features(all_feat)

    source_label = "local cache" if cache_used else "PDF scan"

    # ── Step 3: output ───────────────────────────────────────────────────────
    if args.json:
        result = {
            "module":             args.module,
            "wcn_chips":          sorted(all_wcns),
            "source":             source_label,
            "reference_docs":     doc_names,
            "supported_protocols": protocols,
            "features":           features,
        }
        if module_wifi:
            # Include variant platform data (v3.0) or simplified wifi info (v2.0)
            result["platform"] = {
                "wifi_standard":   module_wifi.get("wifi_standard"),
                "frequency_bands": module_wifi.get("frequency_bands", []),
                "variant_count":   module_wifi.get("variant_count", 0),
                "variants":        module_wifi.get("variants", []),
            }
        # Include doc version metadata from chip cache
        if cache_used and "chips" in cache:
            doc_vers = {}
            for chip in all_wcns:
                key = canonical_chip_key(chip)
                for doc, ver in cache["chips"].get(key, {}).get("doc_versions", {}).items():
                    doc_vers[doc] = ver
            if doc_vers:
                result["doc_versions"] = doc_vers
        if doc_hit_counts:
            result["doc_hit_counts"] = doc_hit_counts
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    sep = "=" * 62
    print(f"\n{sep}")
    print(f"  WiFi Specification Lookup: {args.module}  [{source_label}]")
    print(sep)
    print(f"\n  WCN Chip(s) : {', '.join(sorted(all_wcns))}")
    print(f"  Variants    : {len(set(variants))} found ({', '.join(sorted(set(variants))[:3])}"
          + (" ..." if len(set(variants)) > 3 else "") + ")")

    if module_wifi:
        print(f"\n  WiFi Standard   : {module_wifi.get('wifi_standard', 'N/A')}")
        if module_wifi.get('frequency_bands'):
            print(f"  Freq. Bands     : {', '.join(module_wifi['frequency_bands'])}")
        # v3.0: show variant count and first variant's chipset string
        variant_list = module_wifi.get("variants", [])
        if variant_list:
            print(f"  Platform variants: {module_wifi.get('variant_count', len(variant_list))}")
            # Show chipset of first unique variant
            first_chipset = variant_list[0].get("chipset", "")
            if first_chipset:
                print(f"  Platform chipset: {first_chipset[:80]}"
                      + ("…" if len(first_chipset) > 80 else ""))

    print(f"\n{sep}")
    print(f"  Supported Security Protocols")
    print(sep)
    if protocols:
        for p in protocols:
            print(f"  ✔  {p}")
    else:
        print("  (No protocol information found)")

    if features:
        print(f"\n{sep}")
        print(f"  WiFi Features (chip-level)")
        print(sep)
        CATEGORY_LABELS = {
            "rf_features":      "RF / PHY Features",
            "channel_and_phy":  "Channel & PHY Modes",
            "qos":              "QoS & Traffic Management",
            "roaming":          "Roaming & Mobility",
            "location":         "Location / Positioning",
            "power_save":       "Power Save & Offload",
            "concurrent_modes": "Concurrent Modes",
            "ap_mode":          "AP Mode Features",
            "privacy":          "Privacy & MAC Security",
            "coexistence":      "Coexistence",
        }
        for cat, items in features.items():
            label = CATEGORY_LABELS.get(cat, cat)
            print(f"\n  [{label}]")
            for item in items:
                print(f"    ✔  {item}")

    if doc_names:
        print(f"\n  Reference documents:")
        for name in doc_names:
            hits = doc_hit_counts.get(name)
            prefix = f"[{hits:3d} hits]" if hits is not None else "[cached]"
            print(f"    {prefix}  {name}")
    print()


if __name__ == "__main__":
    main()

